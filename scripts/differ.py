# differ.py
# ----------
#
import boto3
import datetime
import json
import openpyxl
import re
import sys
import time
import yaml

if len(sys.argv) != 3:
    raise Exception('Usage: python3 differ.py [FIRST_XLSX_PATH] [SECOND_XLSX_PATH]')

FILE1=sys.argv[1]
FILE2=sys.argv[2]

def compare_xlsx_files(file1, file2, sheet_name1, sheet_name2, data_row_num, header_row_num):
    # Load both XLSX files
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    # Get the specified sheet from each file
    sheet1 = wb1[sheet_name1]
    sheet2 = wb2[sheet_name2]

    # Get the header row from each sheet
    header_row1 = sheet1[header_row_num]
    header_row2 = sheet2[header_row_num]

    # Check for columns in sheet1 but not sheet2
    for i in range(1, sheet1.max_column + 1):
        header_val = header_row1[i-1].value
        if header_val not in [x.value for x in header_row2]:
            print(f"Column {header_val} is in {file1} but not {file2}")

    # Check for columns in sheet2 but not sheet1
    for i in range(1, sheet2.max_column + 1):
        header_val = header_row2[i-1].value
        if header_val not in [x.value for x in header_row1]:
            print(f"Column {header_val} is in {file2} but not {file1}")

    # Get the row from each sheet
    data_row1 = sheet1[data_row_num]
    data_row2 = sheet2[data_row_num]

    # Compare the values in each cell of the row
    for i in range(len(data_row1)):
        cell1 = data_row1[i].value
        header_value1 = header_row1[i].value
        # find the right column in sheet2
        match = False
        for j in range(len(header_row2)):
            if header_value1 == header_row2[j].value:
                match = True
                cell2 = data_row2[j].value
                if cell1 != cell2:
                    print(f"{header_value1} is different:")
                    print(f"{file1}: {cell1}")
                    print(f"{file2}: {cell2}")
                if match:
                    break
        if not match:
            print(f"ERROR: Could not find header {header_value1} in second spreadsheet!")

compare_xlsx_files(FILE1, FILE2, 'SSLSingleAMIAndCAR', 'SSLSingleAMIAndCARWithContract', 6, 5)
