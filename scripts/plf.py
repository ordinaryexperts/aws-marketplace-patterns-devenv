# plf.py
# ----------
#
import argparse
import boto3
import datetime
import json
import openpyxl
import pystache
import re
import shutil
import sys
import time
import yaml

from oe_patterns_cdk_common.asg import Asg

parser = argparse.ArgumentParser(description='Process some flags and arguments.')

# Optional flags
parser.add_argument('--skip-pricing-update', action='store_true', help='Skip pricing update')
parser.add_argument('--skip-region-update', action='store_true', help='Skip region update')

# Required arguments
parser.add_argument('AMI_ID', help='AMI ID')
parser.add_argument('TEMPLATE_VERSION', help='Template version')

args = parser.parse_args()
skip_pricing_update = args.skip_pricing_update
skip_region_update = args.skip_region_update
AMI = args.AMI_ID
VERSION = args.TEMPLATE_VERSION

OE_MARKUP_PERCENTAGE = 0.10
ANNUAL_SAVINGS_PERCENTAGE = 0.80 # 20% off
MINIMUM_RATE = 0.02
HOURS_IN_A_YEAR = 8760
DEFAULT_REGION = 'us-east-1'

pricing = boto3.client('pricing', region_name=DEFAULT_REGION)

all_regions = {
    'US East (N. Virginia)'     :'us-east-1',
    'US East (Ohio)'            :'us-east-2',
    'US West (N. California)'   :'us-west-1',
    'US West (Oregon)'          :'us-west-2',
    'AWS GovCloud (US-West)'    :'us-gov-west-1',
    'AWS GovCloud (US-East)'    :'us-gov-east-1',
    'Canada (Central)'          :'ca-central-1',
    'EU (Frankfurt)'            :'eu-central-1',
    'EU (Zurich)'               :'eu-central-2',
    'EU (Ireland)'              :'eu-west-1',
    'EU (London)'               :'eu-west-2',
    'EU (Paris)'                :'eu-west-3',
    'EU (Stockholm)'            :'eu-north-1',
    'EU (Milan)'                :'eu-south-1',
    'EU (Spain)'                :'eu-south-2',
    'Africa (Cape Town)'        :'af-south-1',
    'Asia Pacific (Singapore)'  :'ap-southeast-1',
    'Asia Pacific (Sydney)'     :'ap-southeast-2',
    'Asia Pacific (Jakarta)'    :'ap-southeast-3',
    'Asia Pacific (Melbourne)'  :'ap-southeast-4',
    'Asia Pacific (Mumbai)'     :'ap-south-1',
    'Asia Pacific (Hyderabad)'  :'ap-south-2',
    'Asia Pacific (Tokyo)'      :'ap-northeast-1',
    'Asia Pacific (Seoul)'      :'ap-northeast-2',
    'Asia Pacific (Osaka)'      :'ap-northeast-3',
    'Asia Pacific (Hong Kong)'  :'ap-east-1',
    'South America (Sao Paulo)' :'sa-east-1',
    'Middle East (Bahrain)'     :'me-south-1',
    'Middle East (Dubai)'       :'me-south-1',
    'China (Beijing)'           :'none',
    'China (Ningxia)'           :'none',
    'US East (Miami)'           :'none',
    'US West (Los Angeles)'     :'none'
}

plf_config = yaml.load(
    open('/code/plf_config.yaml'),
    Loader=yaml.SafeLoader
)
if plf_config['Architecture'] == 'x86_64':
    allowed_instance_types = Asg.STANDARD_INSTANCE_TYPES
else:
    allowed_instance_types = Asg.GRAVITON_INSTANCE_TYPES

allowed_regions = open('/code/supported_regions.txt').read().split('\n')

def get_highest_hourly_price_for_instance_type(instance_type, allowed_regions):
    highest_hourly_price = 0
    highest_hourly_region = None
    response = pricing.get_products(
        ServiceCode='AmazonEC2',
        Filters = [
            {'Type' :'TERM_MATCH', 'Field':'capacitystatus',  'Value':'Used' },
            {'Type' :'TERM_MATCH', 'Field':'instanceType',    'Value':instance_type },
            {'Type' :'TERM_MATCH', 'Field':'licenseModel',    'Value':'No License required' },
            {'Type' :'TERM_MATCH', 'Field':'operatingSystem', 'Value':'Linux' },
            {'Type' :'TERM_MATCH', 'Field':'preInstalledSw',  'Value':'NA' },
            {'Type' :'TERM_MATCH', 'Field':'tenancy',         'Value':'Shared' },
            {'Type' :'TERM_MATCH', 'Field':'termType',        'Value':'OnDemand' }
        ],
        MaxResults=100
    )
    for price in response['PriceList']:
        priceObj = json.loads(price)
        location = priceObj['product']['attributes']['location']
        if location in all_regions and all_regions[location] in allowed_regions:
            termsKey = next(iter(priceObj['terms']['OnDemand']))
            priceDimensionsKey = next(iter(priceObj['terms']['OnDemand'][termsKey]['priceDimensions']))
            hourly_price = float(priceObj['terms']['OnDemand'][termsKey]['priceDimensions'][priceDimensionsKey]['pricePerUnit']['USD'])
            # print(f'Hourly price for {location} is {hourly_price}')
            if hourly_price > highest_hourly_price:
                highest_hourly_price = hourly_price
                highest_hourly_region = location
    # print(f'Highest price for {instance_type} is ${highest_hourly_price} at {highest_hourly_region}')
    return highest_hourly_price

src = 'plf.xlsx'
now_dt = datetime.datetime.now()
PRODUCT_SLUG = plf_config.get('Product Slug', 'oe')
dst = f"plf-{PRODUCT_SLUG}-v{VERSION.replace('.', '-')}-gen-{now_dt.strftime('%Y%m%d-%H%M%S')}.xlsx"
SHEET_NAME_OPTIONS = ['SSLSingleAMIAndCAR', 'SSLSingleAMIAndCARWithContract']
SHEET_NAME = None

src_wb = openpyxl.load_workbook(src)
for name in SHEET_NAME_OPTIONS:
    if name in src_wb.sheetnames:
        SHEET_NAME = name
        break

if SHEET_NAME is None:
    print("Sheet not found in the source file.")
    exit()

shutil.copyfile(src, dst)

dst_wb = openpyxl.load_workbook(dst)
dst_sheet = dst_wb[SHEET_NAME]

src_sheet = src_wb[SHEET_NAME]
headers = src_sheet[5]
values = src_sheet[src_sheet.max_row]
row_num = src_sheet.max_row + 1

current_column_index = 0
for header in headers:
    column = header.value
    value = ''
    current_value = values[current_column_index].value
    availability_match = re.search(r'(.+) Availability', column)
    if availability_match:
        if skip_region_update:
            value = current_value
        else:
            match_keyword = availability_match.groups()[0]
            # region or instance availability?
            is_instance_match = re.search(r'^(.+)\.(.+)$', match_keyword)
            if is_instance_match:
                if match_keyword in allowed_instance_types:
                    value = 'TRUE'
                else:
                    value = ''
            else:
                if match_keyword in allowed_regions:
                    value = 'TRUE'
                else:
                    value = ''

    price_match = re.search(r'(.+) (Hourly|Annual) Price', column)
    if price_match:
        if skip_pricing_update:
            value = current_value
        else:
            instance_type = price_match.groups()[0]
            if instance_type in allowed_instance_types:
                price_type = price_match.groups()[1]
                price = get_highest_hourly_price_for_instance_type(instance_type, allowed_regions)
                hourly_price_with_markup = round(price * OE_MARKUP_PERCENTAGE, 2)
                if hourly_price_with_markup < MINIMUM_RATE:
                    hourly_price_with_markup = MINIMUM_RATE
                if price_type == 'Hourly':
                    value = '{:.3f}'.format(hourly_price_with_markup)
                else:
                    annual_price = hourly_price_with_markup * HOURS_IN_A_YEAR * ANNUAL_SAVINGS_PERCENTAGE
                    value = '{:.3f}'.format(round(annual_price, 2))
    if not availability_match and not price_match:
        if column in plf_config:
            value = pystache.render(plf_config[column], {'ami': AMI, 'version': VERSION})
    if current_value is None:
        current_value = ''
    if value != current_value:
        print(f"{column} has changed! Old: '{current_value}' New: '{value}'")
        dst_sheet.cell(row=row_num, column=current_column_index+1, value=value)
    else:
        dst_sheet.cell(row=row_num, column=current_column_index+1, value=current_value)
    current_column_index += 1

dst_wb.save(dst)
print(f'PLF saved to {dst}')
